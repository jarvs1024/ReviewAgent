"""周报主入口 — run_weekly_job.

参考 pr-agent: pr_agent/reporting/scheduler.run_weekly_job

流程:
    1. 算 week_start / week_end (默认本周)
    2. 实例化 collector(s) 拉数据 -> sections
    3. 构造 WeeklyArtifact
    4. JSON 落盘
    5. 渲染 markdown
    6. 实例化 notifier 推送 (默认 dry_run)

失败隔离: 任何 collector 抛错都被 scheduler 包成 status='failed' 的 SectionResult;
         notifier 推送失败不会影响 artifact 落盘.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from reviewagent.logging_setup import logger

from .artifact import (
    SCHEMA_VERSION,
    WeeklyArtifact,
    build_artifact,
    iso_week_label,
    write_artifact,
)
from .collectors import Collector, CollectorContext, SectionResult
from .collectors.telemetry import TelemetryCollector
from .config import WeeklyReportConfig
from .notifiers import DeliveryResult, Notifier
from .notifiers.dingtalk import DingTalkNotifier
from .renderer import render_markdown, split_markdown


def _week_bounds(
    now: datetime | None = None,
    tz: timezone | None = None,
    week_offset: int = 0,
) -> tuple[datetime, datetime]:
    """算本周 (或 N 周前) 的 (week_start, week_end).

    week_offset=0 = 本周, -1 = 上周.
    """
    if now is None:
        # 默认 Asia/Shanghai (UTC+8)
        tz = tz or timezone(timedelta(hours=8))
        now = datetime.now(tz)
    if now.tzinfo is None:
        now = now.replace(tzinfo=tz)
    monday = now - timedelta(days=now.weekday())
    monday = monday.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = monday + timedelta(weeks=week_offset)
    week_end = week_start + timedelta(days=7)
    return week_start, week_end


def _build_collector(name: str) -> Collector | None:
    # Local imports 避免启动期循环引用
    from .collectors.merged_mrs import MergedMrsCollector
    table: dict[str, type] = {
        "telemetry": TelemetryCollector,
        "merged_mrs": MergedMrsCollector,
    }
    cls = table.get(name)
    if cls is None:
        logger.warning("reporting.unknown_collector name={}", name)
        return None
    return cls()


def _build_notifier(cfg: WeeklyReportConfig) -> Notifier:
    if cfg.notifier == "dingtalk":
        return DingTalkNotifier(
            webhook_url=cfg.dingtalk_webhook_url,
            secret=cfg.dingtalk_secret,
            retry_attempts=cfg.dingtalk_retry_attempts,
            dry_run=cfg.dingtalk_dry_run,
        )
    raise RuntimeError(f"unsupported notifier: {cfg.notifier!r}")


def run_weekly_job(
    cfg: WeeklyReportConfig | None = None,
    *,
    now: datetime | None = None,
    week_offset: int = 0,
    output_dir: Path | None = None,
    push: bool = False,
) -> dict[str, Any]:
    """跑一次完整周报 job.

    Args:
        cfg: 运行时配置 (None = 从 env 读)
        now: 测试注入 (默认 datetime.now)
        week_offset: 0=本周, -1=上周
        output_dir: artifact/md/xlsx 输出目录 (默认 data/weekly_reports/)
        push: 是否真实推送 (False = 走 notifier 的 dry_run)

    Returns:
        dict 含 artifact_path / markdown_path / xlsx_path / delivery_result
    """
    cfg = cfg or WeeklyReportConfig.from_env()
    output_dir = output_dir or Path("data/weekly_reports")

    tz = timezone(timedelta(hours=8))  # Asia/Shanghai
    week_start, week_end = _week_bounds(now=now, tz=tz, week_offset=week_offset)

    logger.info(
        "reporting.run start week={} ({}) range={} ~ {} push={}",
        iso_week_label(week_start), week_offset, week_start.isoformat(),
        week_end.isoformat(), push,
    )

    # 1. collectors
    ctx = CollectorContext(
        target_project_id=cfg.target_project_id,
        data_dir=str(output_dir),
        timezone=cfg.timezone,
    )
    sections: dict[str, SectionResult] = {}
    for name in cfg.collectors:
        c = _build_collector(name)
        if c is None:
            continue
        try:
            sections[name] = c.collect(week_start=week_start, week_end=week_end, ctx=ctx)
            logger.info("reporting.collector ok name={} status={}", name, sections[name].status)
        except Exception as e:
            logger.exception("reporting.collector crash name={}: {}", name, e)
            sections[name] = SectionResult(status="failed", error=str(e), data=None)

    # 2. artifact
    artifact = build_artifact(
        project_id=cfg.target_project_id or 0,
        week_start=week_start,
        week_end=week_end,
        timezone=cfg.timezone,
        sections=sections,
        report_title=cfg.report_title,
        report_emoji=cfg.report_emoji,
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    artifact_path = output_dir / f"weekly-{artifact.week_label}.json"
    write_artifact(artifact, artifact_path)

    # 3. render md
    md_text = render_markdown(artifact)
    md_path = output_dir / f"weekly-{artifact.week_label}.md"
    md_path.write_text(md_text, encoding="utf-8")
    logger.info("reporting.markdown written path={} bytes={}", md_path, len(md_text.encode("utf-8")))

    # 4. xlsx (复用 weekly_report.py 里的逻辑)
    xlsx_path = output_dir / f"weekly-{artifact.week_label}.xlsx"
    try:
        _render_xlsx(artifact, md_text, xlsx_path)
    except ImportError:
        logger.warning("reporting.xlsx skipped (openpyxl not installed)")
        xlsx_path = None  # type: ignore

    # 5. notify
    delivery: dict[str, Any] = {"skipped": True}
    if push or not cfg.dingtalk_dry_run:
        notifier = _build_notifier(cfg)
        chunks = split_markdown(md_text, chunk_limit=cfg.markdown_chunk_limit)
        result = notifier.send(
            title=f"{cfg.report_emoji} {cfg.report_title} ({artifact.week_label})",
            markdown_chunks=chunks,
        )
        delivery = {
            "notifier": notifier.name,
            "success": result.success,
            "chunks_sent": result.chunks_sent,
            "chunks_total": result.chunks_total,
            "error": result.error,
            "meta": result.meta,
        }
    else:
        # 总是走一次 notifier.send (dry_run), 让 log 看到分片
        notifier = _build_notifier(cfg)
        chunks = split_markdown(md_text, chunk_limit=cfg.markdown_chunk_limit)
        result = notifier.send(
            title=f"{cfg.report_emoji} {cfg.report_title} ({artifact.week_label})",
            markdown_chunks=chunks,
        )
        delivery = {
            "notifier": notifier.name,
            "dry_run": True,
            "chunks_total": result.chunks_total,
        }

    return {
        "week_label": artifact.week_label,
        "artifact_path": str(artifact_path),
        "markdown_path": str(md_path),
        "xlsx_path": str(xlsx_path) if xlsx_path else None,
        "delivery": delivery,
        "sections": {n: sr.status for n, sr in sections.items()},
    }


def _render_xlsx(artifact: WeeklyArtifact, md_text: str, out_path: Path) -> None:
    """生成 xlsx — 复用 weekly_report.py 里的渲染逻辑, 但用 artifact 结构."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: 汇总
    ws = wb.active
    ws.title = "汇总"
    ws.append([f"{artifact.report_emoji} {artifact.report_title}", artifact.week_label])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["周期", f"{artifact.week_start.isoformat()} ~ {artifact.week_end.isoformat()}"])
    ws.append(["Project ID", artifact.project_id])
    ws.append(["生成时间", artifact.generated_at.isoformat() if artifact.generated_at else ""])
    ws.append(["schema_version", SCHEMA_VERSION])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38

    # 后续 sheet 由 sections 拼
    for name, sr in artifact.sections.items():
        if sr.status == "failed" or not sr.data:
            continue
        if name == "telemetry":
            d = sr.data
            # 命令维度
            ws2 = wb.create_sheet("命令维度")
            headers = ["命令", "次数", "成功", "失败", "运行中", "平均耗时(s)", "最长耗时(s)"]
            ws2.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws2.cell(1, c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for cmd, bc in (d.get("by_command") or {}).items():
                ws2.append([cmd, bc["count"], bc["success"], bc["failed"],
                           bc["running"], round(bc["avg_duration_ms"]/1000, 1),
                           round(bc["max_duration_ms"]/1000, 1)])
            for c, w in enumerate([14, 10, 10, 10, 10, 14, 14], 1):
                ws2.column_dimensions[get_column_letter(c)].width = w

            # 日趋势
            ws3 = wb.create_sheet("日趋势")
            ws3.append(["日期", "run 数"])
            ws3.cell(1, 1).font = header_font
            ws3.cell(1, 1).fill = header_fill
            ws3.cell(1, 2).font = header_font
            ws3.cell(1, 2).fill = header_fill
            for day, n in (d.get("by_day") or {}).items():
                ws3.append([day, n])
            ws3.column_dimensions["A"].width = 14
            ws3.column_dimensions["B"].width = 12

            # 活跃 MR
            ws4 = wb.create_sheet("活跃MR")
            headers = ["project_id", "mr_iid", "标题", "作者", "run", "成功", "失败"]
            ws4.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws4.cell(1, c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for mr in (d.get("top_mrs") or [])[:20]:
                ws4.append([mr["project_id"], mr["mr_iid"], mr.get("title", ""),
                           mr.get("author", ""), mr["runs"], mr["success"], mr["failed"]])
            for c, w in enumerate([12, 10, 50, 16, 10, 10, 10], 1):
                ws4.column_dimensions[get_column_letter(c)].width = w

            # 失败 run
            ws5 = wb.create_sheet("失败run")
            headers = ["时间", "命令", "project_id", "mr_iid", "actor", "错误"]
            ws5.append(headers)
            for c in range(1, len(headers) + 1):
                cell = ws5.cell(1, c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for r in (d.get("failed_runs") or [])[:30]:
                ws5.append([r.get("started_at", ""), r.get("command", ""),
                           r.get("project_id", ""), r.get("mr_iid", ""),
                           r.get("actor_username", ""), r.get("error", "")])
            for c, w in enumerate([22, 12, 12, 10, 16, 60], 1):
                ws5.column_dimensions[get_column_letter(c)].width = w

    # Markdown
    ws6 = wb.create_sheet("Markdown")
    ws6["A1"] = md_text
    ws6.column_dimensions["A"].width = 100
    ws6.row_dimensions[1].height = max(15 * md_text.count("\n"), 30)

    wb.save(str(out_path))
    logger.info("reporting.xlsx written path={} bytes={}", out_path, out_path.stat().st_size)


def main() -> int:
    """CLI 入口 — `python -m reviewagent.reporting.runner [args]`"""
    import argparse
    p = argparse.ArgumentParser(description="Run ReviewAgent weekly report")
    p.add_argument("--week-offset", type=int, default=0,
                   help="0=本周, -1=上周")
    p.add_argument("--push", action="store_true",
                   help="真实推送 (默认走 notifier.dry_run=True)")
    p.add_argument("--output-dir", type=Path, default=Path("data/weekly_reports"))
    p.add_argument("--project-id", type=int, default=None,
                   help="覆盖 REVIEWAGENT_WEEKLY_TARGET_PROJECT_ID")
    args = p.parse_args()

    cfg = WeeklyReportConfig.from_env()
    if args.project_id is not None:
        cfg = WeeklyReportConfig(
            **{**cfg.__dict__, "target_project_id": args.project_id}
        )

    result = run_weekly_job(
        cfg=cfg,
        week_offset=args.week_offset,
        output_dir=args.output_dir,
        push=args.push,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
