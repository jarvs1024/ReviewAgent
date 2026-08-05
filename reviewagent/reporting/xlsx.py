"""Weekly report → XLSX (Excel) writer.

Why a separate file:
  - ``artifact.py`` 负责 JSON 落盘; ``renderer.py`` 负责 Markdown.
    XLSX 是另一种消费者, 单独写一个 renderer-style 函数更清晰, 且
    容易在测试里单独构造 fixture 调用.
  - 不引入 xlsxwriter / pandas 等重依赖: 直接用 openpyxl (项目已经依赖).

Sheets produced (随 SectionResult.status 跳过空 section):
  1. "概述"        — 全周整体统计 + section 状态
  2. "本周 MR 检视" — telemetry section 拆出 (top rules / severity / authors)
  3. "合并 MR"      — merged_mrs section 拆出 (iid, title, author, summary_md)
  4. "代码质量扫描" — repo_scan section 拆出 (本周风险模块)

设计原则:
  - 所有 sheet 第一行冻结表头 (便于浏览大数据)
  - 列宽按"该列最长文本"自动扩展, 上限 60
  - 表格带样式: header 加粗 + 浅蓝填充; 数值列右对齐; 文本左对齐
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .artifact import WeeklyArtifact
from .collectors.base import SectionResult


_HEADER_FONT = Font(name="DengXian", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_CELL_FONT = Font(name="DengXian", size=10)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="top")

MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8


def _autofit_columns(ws: Worksheet) -> None:
    """Set column widths based on max string length of each column (max 60 chars)."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            if c.value is None:
                continue
            try:
                length = max(len(line) for line in str(c.value).split("\n"))
            except ValueError:
                length = 0
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, max_len + 2))


def _write_header(ws: Worksheet, headers: Sequence[str]) -> None:
    for idx, name in enumerate(headers, 1):
        c = ws.cell(row=1, column=idx, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _write_rows(ws: Worksheet, rows: Sequence[Sequence[Any]]) -> None:
    for r, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val if val != "" else None)
            cell.font = _CELL_FONT
            if isinstance(val, (int, float)):
                cell.alignment = _RIGHT_ALIGN
            else:
                cell.alignment = _LEFT_ALIGN


def _safe_section(artifact: WeeklyArtifact, name: str) -> SectionResult | None:
    sr = artifact.sections.get(name)
    if sr is None or sr.status != "ok":
        return None
    return sr


def _summary_rows(artifact: WeeklyArtifact) -> list[list[Any]]:
    rows: list[list[Any]] = [
        ["项目 ID", artifact.project_id],
        ["周", artifact.week_label],
        ["数据范围", f"{artifact.week_start.isoformat()[:10]} ~ {artifact.week_end.isoformat()[:10]}"],
        ["时区", artifact.timezone],
        ["生成时间", (artifact.generated_at.isoformat()[:19].replace("T", " ") if artifact.generated_at else "")],
    ]
    for name in ("telemetry", "merged_mrs", "repo_scan"):
        sr = artifact.sections.get(name)
        if sr is None:
            status = "missing"
        else:
            status = sr.status
        rows.append([f"section: {name}", status])
    return rows


def _telemetry_rows(sr: SectionResult) -> list[list[Any]]:
    """Render telemetry section into (key, value) rows for the XLSX."""
    data = sr.data or {}
    rows: list[list[Any]] = []

    def _kv(label: str, key: str) -> None:
        if key in data:
            rows.append([label, data[key]])

    _kv("MR 数 (本周)", "mr_count")
    _kv("累计 MR 数", "mr_total")
    _kv("Suggestion 数 (本周)", "suggestion_count")
    _kv("累计 Suggestion 数", "suggestion_total")
    _kv("累计采纳率", "adoption_rate")
    _kv("平均 duration (ms)", "avg_duration_ms")
    _kv("Run 成功率", "success_rate")

    severity = data.get("severity_breakdown") or {}
    if severity:
        rows.append(["--- severity 分布 ---", ""])
        for sev, cnt in sorted(severity.items(), key=lambda kv: -kv[1]):
            rows.append([f"  severity={sev}", cnt])

    rules = data.get("top_rules") or []
    if rules:
        rows.append(["--- 命中 Top 规则 ---", ""])
        # collector 顺序: top_rules = [[rule_key, count], ...] (tuple 列表)
        for entry in rules[:10]:
            if isinstance(entry, (list, tuple)) and len(entry) >= 2:
                rule_key, count = entry[0], entry[1]
            elif isinstance(entry, dict):
                rule_key, count = entry.get("rule_key", ""), entry.get("count", 0)
            else:
                continue
            rows.append([f"  {rule_key}", f"命中 {count} 次"])

    return rows


def _merged_mrs_rows(sr: SectionResult) -> list[list[Any]]:
    data = sr.data or {}
    mrs = data.get("merged_mrs") or []
    rows: list[list[Any]] = []
    for mr in mrs:
        rows.append([
            mr.get("iid", ""),
            mr.get("title", ""),
            mr.get("author", ""),
            mr.get("files_changed", ""),
            mr.get("lines_added", ""),
            mr.get("lines_deleted", ""),
            (mr.get("summary_md") or "")[:1000],
        ])
    return rows


def _repo_scan_rows(sr: SectionResult) -> list[list[Any]]:
    data = sr.data or {}
    rows: list[list[Any]] = []
    hot = data.get("hot_modules") or []
    for m in hot:
        rows.append([
            m.get("module", ""),
            m.get("files", ""),
            m.get("lines_changed", ""),
            m.get("risk", ""),
            (m.get("why") or "")[:800],
        ])
    return rows


def write_xlsx(artifact: WeeklyArtifact, out_path: Path) -> Path:
    """渲染 WeeklyArtifact 到 XLSX 文件, 返回 out_path.

    Sheets:
      1. "概述"        — 总览 + section 状态 (固定列 2)
      2. "本周 MR 检视" — 当 telemetry 存在时才有
      3. "合并 MR"      — 当 merged_mrs 存在时才有; 列 iid/title/author/files/+/−/summary
      4. "代码质量扫描" — 当 repo_scan 存在时才有
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 删除默认 Sheet, 改名为 "概述"
    summary_ws = wb.active
    summary_ws.title = "概述"
    _write_header(summary_ws, ["指标", "值"])
    _write_rows(summary_ws, _summary_rows(artifact))
    summary_ws.column_dimensions["A"].width = 26
    summary_ws.column_dimensions["B"].width = 60

    telemetry = _safe_section(artifact, "telemetry")
    if telemetry is not None:
        rows = _telemetry_rows(telemetry)
        if rows:
            ws = wb.create_sheet("本周 MR 检视")
            _write_header(ws, ["指标", "值"])
            _write_rows(ws, rows)
            _autofit_columns(ws)

    merged = _safe_section(artifact, "merged_mrs")
    if merged is not None:
        ws = wb.create_sheet("合并 MR")
        _write_header(ws, ["MR iid", "标题", "作者", "改文件数", "新增行", "删除行", "摘要"])
        rows = _merged_mrs_rows(merged)
        if rows:
            _write_rows(ws, rows)
        else:
            _write_rows(ws, [["(本周无合并 MR)"]])
        _autofit_columns(ws)

    repo_scan = _safe_section(artifact, "repo_scan")
    if repo_scan is not None:
        ws = wb.create_sheet("代码质量扫描")
        _write_header(ws, ["模块", "文件数", "总变更", "风险", "原因"])
        rows = _repo_scan_rows(repo_scan)
        if rows:
            _write_rows(ws, rows)
        else:
            _write_rows(ws, [["(本周无变更)"]])
        _autofit_columns(ws)

    wb.save(out_path)
    return out_path


__all__ = ["write_xlsx"]
