"""Weekly report → XLSX renderer tests."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

from reviewagent.reporting.collectors.base import SectionResult
from reviewagent.reporting.artifact import WeeklyArtifact
from reviewagent.reporting.xlsx import write_xlsx


def _make_artifact(tmp_path: Path, *, with_telemetry=True, with_merged=True, with_scan=True) -> WeeklyArtifact:
    sections = {}
    if with_telemetry:
        sections["telemetry"] = SectionResult(
            status="ok",
            data={
                "mr_count": 2,
                "mr_total": 40,
                "suggestion_count": 8,
                "suggestion_total": 386,
                "adoption_rate": 0.158,
                "avg_duration_ms": 12345,
                "severity_breakdown": {"high": 4, "medium": 3, "low": 1},
                "top_rules": [
                    ["SSD-RULE-NO-MUTABLE-DEFAULT", 4],
                    ["SSD-RULE-NO-LOG-EXC", 3],
                ],
            },
        )
    if with_merged:
        sections["merged_mrs"] = SectionResult(
            status="ok",
            data={
                "merged_mrs": [
                    {
                        "iid": 181,
                        "title": "新增 buggy_module 故意违规样例",
                        "author": "tester",
                        "files_changed": 1,
                        "lines_added": 50,
                        "lines_deleted": 0,
                        "summary_md": "故意违规样例覆盖多条 SSD 规则",
                    },
                ],
            },
        )
    if with_scan:
        sections["repo_scan"] = SectionResult(
            status="ok",
            data={
                "hot_modules": [
                    {"module": "services", "files": 2, "lines_changed": 50, "risk": "high", "why": "频繁变更"},
                ],
            },
        )
    tz = timezone(timedelta(hours=8))
    return WeeklyArtifact(
        project_id=34,
        week_label="2026-W32",
        week_start=datetime(2026, 8, 3, tzinfo=tz),
        week_end=datetime(2026, 8, 10, tzinfo=tz),
        timezone="Asia/Shanghai",
        sections=sections,
        generated_at=datetime(2026, 8, 3, 12, 0, tzinfo=tz),
    )


def test_write_xlsx_creates_file(tmp_path: Path) -> None:
    """xlsx 文件正常生成, 4 个 sheet (概述 + 3 section), 全部可打开."""
    artifact = _make_artifact(tmp_path)
    out = tmp_path / "weekly-test.xlsx"
    p = write_xlsx(artifact, out)
    assert p == out
    assert out.exists()
    assert out.stat().st_size > 1024  # 至少 1 KB

    # 用 openpyxl 重新打开验证
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == ["概述", "本周 MR 检视", "合并 MR", "代码质量扫描"]

    summary_ws = wb["概述"]
    # row 1: header (指标 / 值); row 2 开始: 数据
    assert summary_ws.cell(row=1, column=1).value == "指标"
    assert summary_ws.cell(row=1, column=2).value == "值"
    assert summary_ws.cell(row=2, column=1).value == "项目 ID"
    assert summary_ws.cell(row=2, column=2).value == 34

    tel_ws = wb["本周 MR 检视"]
    assert tel_ws.cell(row=1, column=1).value == "指标"
    assert tel_ws.cell(row=1, column=2).value == "值"
    assert tel_ws.cell(row=2, column=1).value == "MR 数 (本周)"
    assert tel_ws.cell(row=2, column=2).value == 2

    merged_ws = wb["合并 MR"]
    assert merged_ws.cell(row=1, column=1).value == "MR iid"
    # iid 181 在 row 2
    assert merged_ws.cell(row=2, column=1).value == 181

    scan_ws = wb["代码质量扫描"]
    assert scan_ws.cell(row=1, column=1).value == "模块"
    assert "services" in str(scan_ws.cell(row=2, column=1).value or "")


def test_write_xlsx_missing_sections(tmp_path: Path) -> None:
    """section 缺失或 failed 时, 对应 sheet 不生成 (而非空)."""
    artifact = _make_artifact(tmp_path, with_telemetry=False)
    out = tmp_path / "weekly-t.xlsx"
    write_xlsx(artifact, out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    # 只有概述 + 不带 telemetry 的两个 section
    assert "概述" in wb.sheetnames
    assert "本周 MR 检视" not in wb.sheetnames
    assert "合并 MR" in wb.sheetnames
    assert "代码质量扫描" in wb.sheetnames


def test_write_xlsx_empty_week(tmp_path: Path) -> None:
    """section 全空时, 合并 MR / 代码质量扫描 表里有提示行但不崩."""
    artifact = _make_artifact(tmp_path, with_telemetry=True, with_merged=True, with_scan=True)
    # 清空 merged_mrs + repo_scan data
    artifact.sections["merged_mrs"].data = {"merged_mrs": []}
    artifact.sections["repo_scan"].data = {"hot_modules": []}

    out = tmp_path / "weekly-empty.xlsx"
    write_xlsx(artifact, out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    merged_ws = wb["合并 MR"]
    # 第 1 行是 header, 第 2 行是 "(本周无合并 MR)"
    assert "(本周无合并 MR)" in str(merged_ws.cell(row=2, column=1).value or "")


def test_write_xlsx_creates_parent_dirs(tmp_path: Path) -> None:
    """深层 output_dir 自动 mkdir -p."""
    nested = tmp_path / "a" / "b" / "c"
    artifact = _make_artifact(nested)
    write_xlsx(artifact, nested / "weekly.xlsx")
    assert (nested / "weekly.xlsx").exists()
