"""ReviewAgent 周报生成 — 从 telemetry 读数据, 出 Markdown + xlsx.

用法:
    python scripts/weekly_report.py                       # 默认本周 (周一到周日)
    python scripts/weekly_report.py --since 2026-07-21    # 自定义起始
    python scripts/weekly_report.py --until 2026-07-28
    python scripts/weekly_report.py --no-xlsx             # 只出 markdown

输出:
    data/weekly_reports/weekly-YYYY-MM-DD.md
    data/weekly_reports/weekly-YYYY-MM-DD.xlsx

数据源: reviewagent.data.telemetry.db (SQLite)
       也支持通过 reviewagent API (--from-api 模式)
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

# 允许从仓库根直接 import
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from reviewagent.config import config  # noqa: E402
from reviewagent.logging_setup import logger, setup_logging  # noqa: E402

DEFAULT_OUTPUT_DIR = REPO_ROOT / "data" / "weekly_reports"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--since", help="起始日期 (ISO 8601 date 或 datetime), 默认本周一")
    p.add_argument("--until", help="结束日期 (exclusive), 默认下周一")
    p.add_argument("--no-xlsx", action="store_true", help="不生成 xlsx")
    p.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR,
                   help=f"输出目录 (默认: {DEFAULT_OUTPUT_DIR})")
    p.add_argument("--db", type=Path, default=None, help="telemetry db 路径 (默认读 config)")
    return p.parse_args()


def resolve_range(since: str | None, until: str | None) -> tuple[datetime, datetime]:
    """确定周报时间窗口 (UTC). 默认本周 (周一 00:00:00 UTC ~ 下周一 00:00:00 UTC)."""
    today = date.today()
    if since:
        s = date.fromisoformat(since)
    else:
        # 本周一
        s = today - timedelta(days=today.weekday())
    if until:
        u = date.fromisoformat(until)
    else:
        u = s + timedelta(days=7)
    start = datetime(s.year, s.month, s.day, tzinfo=timezone.utc)
    end = datetime(u.year, u.month, u.day, tzinfo=timezone.utc)
    return start, end


def load_runs(db_path: Path, since: datetime, until: datetime) -> list[dict]:
    """从 SQLite 拉 runs + 关联 MR 标题."""
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT r.*, m.title as mr_title, m.author_sticky as mr_author
        FROM review_runs r
        LEFT JOIN mr_activity m
          ON m.project_id = r.project_id AND m.mr_iid = r.mr_iid
        WHERE r.started_at >= ? AND r.started_at < ?
        ORDER BY r.started_at DESC
        """,
        (since.isoformat(), until.isoformat()),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def aggregate(runs: list[dict]) -> dict:
    """聚合统计 — 多维度."""
    total = len(runs)
    by_command: dict[str, dict] = {}
    by_status: dict[str, int] = defaultdict(int)
    by_day: dict[str, int] = defaultdict(int)
    by_mr: dict[tuple[int, int], dict] = {}
    durations_by_command: dict[str, list[int]] = defaultdict(list)
    failed_runs: list[dict] = []

    for r in runs:
        cmd = r["command"]
        st = r["status"]
        dur = r["duration_ms"] or 0
        day = (r["started_at"] or "")[:10]
        bc = by_command.setdefault(cmd, {
            "count": 0, "success": 0, "failed": 0, "timeout": 0, "running": 0,
            "total_duration_ms": 0, "max_duration_ms": 0,
        })
        bc["count"] += 1
        if st in bc:
            bc[st] += 1
        bc["total_duration_ms"] += dur
        bc["max_duration_ms"] = max(bc["max_duration_ms"], dur)
        by_status[st] += 1
        by_day[day] += 1
        durations_by_command[cmd].append(dur)
        mr_key = (r["project_id"], r["mr_iid"])
        bm = by_mr.setdefault(mr_key, {
            "title": r.get("mr_title") or "?",
            "author": r.get("mr_author") or "?",
            "runs": 0, "success": 0, "failed": 0,
        })
        bm["runs"] += 1
        if st == "success":
            bm["success"] += 1
        elif st in ("failed", "timeout"):
            bm["failed"] += 1
        if st in ("failed", "timeout"):
            failed_runs.append(r)

    # 计算 command 级 avg
    for cmd, bc in by_command.items():
        if bc["count"]:
            bc["avg_duration_ms"] = int(bc["total_duration_ms"] / bc["count"])
        else:
            bc["avg_duration_ms"] = 0

    success = by_status.get("success", 0)
    fail = by_status.get("failed", 0) + by_status.get("timeout", 0)
    success_rate = (success / total * 100) if total else 0.0
    avg_duration = (
        int(sum(r["duration_ms"] or 0 for r in runs) / total) if total else 0
    )

    # 排序
    top_mrs = sorted(
        by_mr.items(),
        key=lambda kv: (kv[1]["runs"], kv[1]["failed"]),
        reverse=True,
    )

    return {
        "total": total,
        "success": success,
        "failed": fail,
        "running": by_status.get("running", 0),
        "success_rate": round(success_rate, 1),
        "avg_duration_ms": avg_duration,
        "by_command": by_command,
        "by_status": dict(by_status),
        "by_day": dict(sorted(by_day.items())),
        "top_mrs": [{"project_id": k[0], "mr_iid": k[1], **v} for k, v in top_mrs],
        "failed_runs": failed_runs,
    }


def render_markdown(stats: dict, since: datetime, until: datetime) -> str:
    """生成 markdown 周报."""
    s_str = since.strftime("%Y-%m-%d")
    u_str = (until - timedelta(days=1)).strftime("%Y-%m-%d")
    lines: list[str] = []
    lines.append(f"# ReviewAgent 周报 ({s_str} ~ {u_str})")
    lines.append("")
    lines.append("> 本周报由 `scripts/weekly_report.py` 自动生成，源数据: telemetry.db")
    lines.append("")

    # 1. 总览
    lines.append("## 总览")
    lines.append("")
    lines.append(f"- **总 run 数**: {stats['total']}")
    lines.append(f"- **成功**: {stats['success']}  •  **失败/超时**: {stats['failed']}  •  **运行中**: {stats['running']}")
    lines.append(f"- **成功率**: {stats['success_rate']}%")
    lines.append(f"- **平均耗时**: {stats['avg_duration_ms'] / 1000:.1f}s")
    lines.append("")

    # 2. 按命令维度
    lines.append("## 按命令维度")
    lines.append("")
    lines.append("| 命令 | 次数 | 成功 | 失败 | 运行中 | 平均耗时 | 最长耗时 |")
    lines.append("|---|---|---|---|---|---|---|")
    for cmd, bc in stats["by_command"].items():
        lines.append(
            f"| `{cmd}` | {bc['count']} | {bc['success']} | {bc['failed']} | {bc['running']} | "
            f"{bc['avg_duration_ms']/1000:.1f}s | {bc['max_duration_ms']/1000:.1f}s |"
        )
    lines.append("")

    # 3. 按状态
    lines.append("## 按状态分布")
    lines.append("")
    lines.append("| 状态 | 数量 |")
    lines.append("|---|---|")
    for st, n in stats["by_status"].items():
        lines.append(f"| `{st}` | {n} |")
    lines.append("")

    # 4. 按日趋势
    if stats["by_day"]:
        lines.append("## 按日趋势")
        lines.append("")
        lines.append("| 日期 | run 数 |")
        lines.append("|---|---|")
        for day, n in stats["by_day"].items():
            bar = "█" * min(n, 30)
            lines.append(f"| {day} | {n} {bar} |")
        lines.append("")

    # 5. Top MRs
    if stats["top_mrs"]:
        lines.append("## 活跃 MR Top 10")
        lines.append("")
        lines.append("| Project | MR | 标题 | 作者 | run | 成功 | 失败 |")
        lines.append("|---|---|---|---|---|---|---|")
        for mr in stats["top_mrs"][:10]:
            title = (mr["title"] or "").replace("|", "/")
            if len(title) > 50:
                title = title[:48] + "…"
            lines.append(
                f"| {mr['project_id']} | !{mr['mr_iid']} | {title} | `{mr['author']}` | "
                f"{mr['runs']} | {mr['success']} | {mr['failed']} |"
            )
        lines.append("")

    # 6. 失败 run
    if stats["failed_runs"]:
        lines.append("## 失败 run 详情")
        lines.append("")
        lines.append("| 时间 | 命令 | Project!MR | actor | 错误片段 |")
        lines.append("|---|---|---|---|---|")
        for r in stats["failed_runs"][:20]:
            t = (r["started_at"] or "")[:19]
            err = (r.get("error") or "")[:60].replace("|", "/").replace("\n", " ")
            lines.append(
                f"| {t} | `{r['command']}` | {r['project_id']}!{r['mr_iid']} | "
                f"`{r.get('actor_username','')}` | {err} |"
            )
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append(f"_生成时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}_")
    return "\n".join(lines)


def render_xlsx(stats: dict, since: datetime, until: datetime, md_text: str, out_path: Path) -> None:
    """生成 xlsx — 4 sheet: 汇总 / 命令维度 / 日趋势 / Top MRs / 失败 run / Markdown."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # ---- Sheet 1: 汇总 ----
    ws = wb.active
    ws.title = "汇总"
    s_str = since.strftime("%Y-%m-%d")
    u_str = (until - timedelta(days=1)).strftime("%Y-%m-%d")
    ws.append(["ReviewAgent 周报", f"{s_str} ~ {u_str}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    summary_rows = [
        ("总 run 数", stats["total"]),
        ("成功", stats["success"]),
        ("失败/超时", stats["failed"]),
        ("运行中", stats["running"]),
        ("成功率", f"{stats['success_rate']}%"),
        ("平均耗时", f"{stats['avg_duration_ms']/1000:.1f}s"),
    ]
    for k, v in summary_rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    # ---- Sheet 2: 命令维度 ----
    ws = wb.create_sheet("命令维度")
    headers = ["命令", "次数", "成功", "失败", "运行中", "平均耗时(s)", "最长耗时(s)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for cmd, bc in stats["by_command"].items():
        ws.append([
            cmd, bc["count"], bc["success"], bc["failed"], bc["running"],
            round(bc["avg_duration_ms"]/1000, 1),
            round(bc["max_duration_ms"]/1000, 1),
        ])
    for c, w in enumerate([14, 10, 10, 10, 10, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet 3: 日趋势 ----
    ws = wb.create_sheet("日趋势")
    ws.append(["日期", "run 数"])
    for c in range(1, 3):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for day, n in stats["by_day"].items():
        ws.append([day, n])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12

    # ---- Sheet 4: Top MRs ----
    ws = wb.create_sheet("活跃MR")
    headers = ["project_id", "mr_iid", "标题", "作者", "run 数", "成功", "失败"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for mr in stats["top_mrs"][:30]:
        ws.append([
            mr["project_id"], mr["mr_iid"], mr.get("title", ""),
            mr.get("author", ""), mr["runs"], mr["success"], mr["failed"],
        ])
    for c, w in enumerate([12, 10, 50, 16, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet 5: 失败 run ----
    ws = wb.create_sheet("失败run")
    headers = ["时间", "命令", "project_id", "mr_iid", "actor", "错误"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for r in stats["failed_runs"][:50]:
        ws.append([
            r.get("started_at", ""), r["command"],
            r["project_id"], r["mr_iid"],
            r.get("actor_username", ""), r.get("error", ""),
        ])
    for c, w in enumerate([22, 12, 12, 10, 16, 60], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet 6: Markdown ----
    ws = wb.create_sheet("Markdown")
    ws["A1"] = md_text
    ws.column_dimensions["A"].width = 100
    ws.row_dimensions[1].height = max(15 * md_text.count("\n"), 30)

    wb.save(str(out_path))


def main() -> int:
    setup_logging()
    args = parse_args()
    since, until = resolve_range(args.since, args.until)
    logger.info("weekly_report range: {} ~ {}", since.isoformat(), until.isoformat())

    db_path = args.db or Path(config.sqlite_path)
    if not db_path.exists():
        print(f"[FAIL] telemetry db not found: {db_path}", file=sys.stderr)
        return 1

    runs = load_runs(db_path, since, until)
    stats = aggregate(runs)
    md = render_markdown(stats, since, until)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    s_str = since.strftime("%Y-%m-%d")
    md_path = args.output_dir / f"weekly-{s_str}.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"[ok] markdown -> {md_path} ({len(md)} bytes)")

    if not args.no_xlsx:
        xlsx_path = args.output_dir / f"weekly-{s_str}.xlsx"
        render_xlsx(stats, since, until, md, xlsx_path)
        print(f"[ok] xlsx     -> {xlsx_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
